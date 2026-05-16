from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Avg, Q, F, Sum, DurationField
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
from .models import OutageReport, ReportUpdate
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@staff_member_required
def analytics_dashboard(request):
    """Main analytics dashboard view"""
    return render(request, 'admin/analytics/dashboard.html')

@staff_member_required
def api_analytics_overview(request):
    """API endpoint for overview metrics"""
    # Date range filter
    days = request.GET.get('days', 30)
    start_date = timezone.now() - timedelta(days=int(days))
    
    reports = OutageReport.objects.filter(reported_at__gte=start_date)
    
    # Calculate metrics
    total_reports = reports.count()
    active_reports = reports.exclude(status='resolved').count()
    resolved_reports = reports.filter(status='resolved').count()
    
    # Average resolution time (in hours)
    resolved_reports_with_time = reports.filter(
        status='resolved', 
        resolved_at__isnull=False
    )
    
    avg_resolution_hours = None
    if resolved_reports_with_time.exists():
        total_hours = 0
        for report in resolved_reports_with_time:
            delta = report.resolved_at - report.reported_at
            total_hours += delta.total_seconds() / 3600
        avg_resolution_hours = round(total_hours / resolved_reports_with_time.count(), 1)
    
    # Resolution rate
    resolution_rate = round((resolved_reports / total_reports * 100), 1) if total_reports > 0 else 0
    
    # Reports by type
    reports_by_type = list(reports.values('outage_type').annotate(
        count=Count('id')
    ).values('outage_type', 'count'))
    
    # Reports by status
    reports_by_status = list(reports.values('status').annotate(
        count=Count('id')
    ).values('status', 'count'))
    
    # Most affected villages
    top_villages = list(reports.values('location_text').annotate(
        count=Count('id')
    ).order_by('-count')[:10])
    
    data = {
        'total_reports': total_reports,
        'active_reports': active_reports,
        'resolved_reports': resolved_reports,
        'avg_resolution_hours': avg_resolution_hours,
        'resolution_rate': resolution_rate,
        'reports_by_type': reports_by_type,
        'reports_by_status': reports_by_status,
        'top_villages': top_villages,
    }
    
    return JsonResponse({'success': True, 'data': data})

@staff_member_required
def api_analytics_trends(request):
    """API endpoint for trend data (daily/weekly/monthly)"""
    period = request.GET.get('period', 'weekly')  # daily, weekly, monthly
    days = request.GET.get('days', 90)
    start_date = timezone.now() - timedelta(days=int(days))
    
    reports = OutageReport.objects.filter(reported_at__gte=start_date)
    
    if period == 'daily':
        trend_data = reports.annotate(
            date=TruncDate('reported_at')
        ).values('date').annotate(
            count=Count('id'),
            resolved=Count('id', filter=Q(status='resolved'))
        ).order_by('date')
        
        labels = [item['date'].strftime('%b %d') for item in trend_data]
        counts = [item['count'] for item in trend_data]
        resolved = [item['resolved'] for item in trend_data]
        
    elif period == 'weekly':
        trend_data = reports.annotate(
            week=TruncWeek('reported_at')
        ).values('week').annotate(
            count=Count('id'),
            resolved=Count('id', filter=Q(status='resolved'))
        ).order_by('week')
        
        labels = [f"Week {i+1}" for i in range(len(trend_data))]
        counts = [item['count'] for item in trend_data]
        resolved = [item['resolved'] for item in trend_data]
        
    else:  # monthly
        trend_data = reports.annotate(
            month=TruncMonth('reported_at')
        ).values('month').annotate(
            count=Count('id'),
            resolved=Count('id', filter=Q(status='resolved'))
        ).order_by('month')
        
        labels = [item['month'].strftime('%b %Y') for item in trend_data]
        counts = [item['count'] for item in trend_data]
        resolved = [item['resolved'] for item in trend_data]
    
    return JsonResponse({
        'success': True,
        'data': {
            'labels': labels,
            'reports': counts,
            'resolved': resolved,
            'period': period
        }
    })

@staff_member_required
def api_analytics_response_time(request):
    """API endpoint for response time analytics"""
    days = request.GET.get('days', 90)
    start_date = timezone.now() - timedelta(days=int(days))
    
    # Response time by outage type
    reports = OutageReport.objects.filter(
        status='resolved',
        resolved_at__isnull=False,
        reported_at__gte=start_date
    )
    
    response_by_type = []
    for outage_type in ['electricity', 'water', 'network']:
        type_reports = reports.filter(outage_type=outage_type)
        if type_reports.exists():
            avg_hours = 0
            for report in type_reports:
                delta = report.resolved_at - report.reported_at
                avg_hours += delta.total_seconds() / 3600
            avg_hours = round(avg_hours / type_reports.count(), 1)
            
            response_by_type.append({
                'type': outage_type,
                'type_display': dict(OutageReport.OUTAGE_TYPES).get(outage_type, outage_type),
                'avg_hours': avg_hours,
                'count': type_reports.count()
            })
    
    # Response time trend (weekly averages)
    weekly_data = []
    for i in range(8):  # Last 8 weeks
        week_start = timezone.now() - timedelta(weeks=i+1)
        week_end = timezone.now() - timedelta(weeks=i)
        
        week_reports = reports.filter(
            reported_at__gte=week_start,
            reported_at__lt=week_end
        )
        
        if week_reports.exists():
            total_hours = 0
            for report in week_reports:
                delta = report.resolved_at - report.reported_at
                total_hours += delta.total_seconds() / 3600
            avg_hours = round(total_hours / week_reports.count(), 1)
        else:
            avg_hours = 0
        
        weekly_data.append({
            'week': f"Week {8-i}",
            'avg_hours': avg_hours,
            'count': week_reports.count()
        })
    
    return JsonResponse({
        'success': True,
        'data': {
            'by_type': response_by_type,
            'weekly_trend': weekly_data[::-1]  # Reverse for chronological order
        }
    })

@staff_member_required
def api_analytics_heatmap_data(request):
    """API endpoint for geographical heatmap data"""
    days = request.GET.get('days', 30)
    start_date = timezone.now() - timedelta(days=int(days))
    
    reports = OutageReport.objects.filter(
        reported_at__gte=start_date,
        latitude__isnull=False,
        longitude__isnull=False
    )
    
    heatmap_points = []
    for report in reports:
        # Weight based on status and type
        weight = 1.0
        if report.status == 'pending':
            weight = 1.5
        elif report.status == 'verified':
            weight = 1.2
        elif report.status == 'resolved':
            weight = 0.3
        
        if report.outage_type == 'electricity':
            weight *= 1.2  # Electricity issues get higher weight
        
        heatmap_points.append({
            'lat': float(report.latitude),
            'lng': float(report.longitude),
            'weight': weight
        })
    
    return JsonResponse({
        'success': True,
        'data': heatmap_points
    })

@staff_member_required
def api_analytics_performance(request):
    """API endpoint for team performance metrics"""
    days = request.GET.get('days', 30)
    start_date = timezone.now() - timedelta(days=int(days))
    
    # Reports resolved per day (last 30 days)
    daily_resolved = []
    for i in range(30):
        day = timezone.now() - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        resolved_count = OutageReport.objects.filter(
            status='resolved',
            resolved_at__gte=day_start,
            resolved_at__lt=day_end
        ).count()
        
        daily_resolved.append({
            'date': day.strftime('%b %d'),
            'count': resolved_count
        })
    
    # Average time to first verification
    verified_reports = OutageReport.objects.filter(
        verified_at__isnull=False,
        reported_at__gte=start_date
    )
    
    avg_verification_hours = None
    if verified_reports.exists():
        total_hours = 0
        for report in verified_reports:
            delta = report.verified_at - report.reported_at
            total_hours += delta.total_seconds() / 3600
        avg_verification_hours = round(total_hours / verified_reports.count(), 1)
    
    # Reports by hour of day (for identifying peak reporting times)
    hour_distribution = []
    for hour in range(24):
        count = OutageReport.objects.filter(
            reported_at__hour=hour,
            reported_at__gte=start_date
        ).count()
        hour_distribution.append({
            'hour': hour,
            'count': count,
            'label': f"{hour}:00" if hour != 0 else "Midnight"
        })
    
    return JsonResponse({
        'success': True,
        'data': {
            'daily_resolved': daily_resolved[::-1],  # Reverse for chronological
            'avg_verification_hours': avg_verification_hours,
            'hour_distribution': hour_distribution
        }
    })

@staff_member_required
def api_analytics_export_csv(request):
    """Export reports data to CSV"""
    report_type = request.GET.get('type', 'all')
    days = request.GET.get('days', 90)
    start_date = timezone.now() - timedelta(days=int(days))
    
    reports = OutageReport.objects.filter(reported_at__gte=start_date)
    
    if report_type == 'resolved':
        reports = reports.filter(status='resolved')
    elif report_type == 'active':
        reports = reports.exclude(status='resolved')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="gridwatch_reports_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Report ID', 'Type', 'Location', 'Description', 'Status', 
        'Reported By', 'Contact', 'Reported At', 'Resolved At',
        'Resolution Time (Hours)', 'Latitude', 'Longitude'
    ])
    
    for report in reports:
        resolution_time = None
        if report.resolved_at:
            delta = report.resolved_at - report.reported_at
            resolution_time = round(delta.total_seconds() / 3600, 1)
        
        writer.writerow([
            report.report_id,
            report.get_outage_type_display(),
            report.location_text,
            report.description[:200],
            report.get_status_display(),
            report.user.email if report.user else 'Anonymous',
            report.contact_info or '',
            report.reported_at.strftime('%Y-%m-%d %H:%M:%S'),
            report.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if report.resolved_at else '',
            resolution_time or '',
            report.latitude or '',
            report.longitude or '',
        ])
    
    return response

@staff_member_required
def api_analytics_export_excel(request):
    """Export reports data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    report_type = request.GET.get('type', 'all')
    days = request.GET.get('days', 90)
    start_date = timezone.now() - timedelta(days=int(days))
    
    reports = OutageReport.objects.filter(reported_at__gte=start_date)
    
    if report_type == 'resolved':
        reports = reports.filter(status='resolved')
    elif report_type == 'active':
        reports = reports.exclude(status='resolved')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "GridWatch Reports"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'Report ID', 'Type', 'Location', 'Description', 'Status',
        'Reported By', 'Contact', 'Reported At', 'Resolved At',
        'Resolution Time (Hours)', 'Latitude', 'Longitude'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, report in enumerate(reports, 2):
        resolution_time = None
        if report.resolved_at:
            delta = report.resolved_at - report.reported_at
            resolution_time = round(delta.total_seconds() / 3600, 1)
        
        ws.cell(row=row, column=1, value=report.report_id)
        ws.cell(row=row, column=2, value=report.get_outage_type_display())
        ws.cell(row=row, column=3, value=report.location_text)
        ws.cell(row=row, column=4, value=report.description[:200])
        ws.cell(row=row, column=5, value=report.get_status_display())
        ws.cell(row=row, column=6, value=report.user.email if report.user else 'Anonymous')
        ws.cell(row=row, column=7, value=report.contact_info or '')
        ws.cell(row=row, column=8, value=report.reported_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=9, value=report.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if report.resolved_at else '')
        ws.cell(row=row, column=10, value=resolution_time or '')
        ws.cell(row=row, column=11, value=str(report.latitude or ''))
        ws.cell(row=row, column=12, value=str(report.longitude or ''))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gridwatch_reports_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@staff_member_required
def api_analytics_summary_pdf(request):
    """Generate PDF summary report (using reportlab)"""
    # Note: Requires reportlab package
    # pip install reportlab
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # This is a simplified version - expand as needed
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="gridwatch_summary_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2E7D32'),
        spaceAfter=30
    )
    story.append(Paragraph("GridWatch Analytics Report", title_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 24))
    
    # Summary metrics
    total_reports = OutageReport.objects.count()
    active_reports = OutageReport.objects.exclude(status='resolved').count()
    resolved_reports = OutageReport.objects.filter(status='resolved').count()
    
    metrics_data = [
        ['Total Reports', str(total_reports)],
        ['Active Reports', str(active_reports)],
        ['Resolved Reports', str(resolved_reports)],
        ['Resolution Rate', f"{round(resolved_reports/total_reports*100, 1)}%" if total_reports > 0 else "0%"]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[2*inch, 2*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(metrics_table)
    
    doc.build(story)
    return response